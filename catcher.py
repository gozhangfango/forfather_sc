import os
import shutil                       # 这个模块负责做一些目录操作
import xml.etree.ElementTree as ET  # 这个模块负责解析xml配置文件
import datetime

import PIL.ImageGrab                                      # 这个模块负责屏幕截图
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.schedulers.background import BackgroundScheduler
import pytesseract
from openpyxl import Workbook,load_workbook               # 这个模块负责操作excel文件
import threading
import matplotlib.pyplot as plt
from matplotlib import animation
import numpy as np


def main():
    global pre_date,rectlist,datalist
    print("程序启动")

    # 第一部分，读取配置文件
    tree = ET.ElementTree(file='config.xml')
    root = tree.getroot()
    rectlist = []
    displaypar = 1
    step = 10
    displaypar = float(tree.find("display-par").text)
    step = tree.find("step").text
    ylimit_min = int(tree.find('ylimit_min').text)
    ylimit_max = int(tree.find('ylimit_max').text)
    for elem in tree.iter(tag='rect'):
        tempitem = RectItem()
        tempitem.name = elem.get("name")
        tempitem.topleftwidth = int(elem.find("top-left-dot").get("width")) * displaypar
        tempitem.topleftheight = int(elem.find("top-left-dot").get("height")) * displaypar
        tempitem.bottomrightwidth = int(elem.find("bottom-right-dot").get("width")) * displaypar
        tempitem.bottomrightheight = int(elem.find("bottom-right-dot").get("height")) * displaypar
        rectlist.append(tempitem)
    print("解析配置文件成功，您要分析的区域是：")
    for item in rectlist:
        print(item)
    print("您的显示参数是：" + str(displaypar))
    print("记录步长设置为：" + str(step))

    # 第二部分，初始化前值列表
    for rect in rectlist:
        pre_date.append(0)
        datalist.append([])

    # 第三部分，检测文件是否存在，没有则创建，并初始化表头
    if os.path.exists(res_path):
        print('检测到文件已经存在，将追加写入')
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = '监控数据'
        head_list = (['时间'])
        for rect in rectlist:
            head_list.append(rect.name)
        ws1.append(head_list)
        wb.save(res_path)

    # 第四部分，开始按照时间规则执行抓取任务
    par_str = r'*/' + step
    sched = BackgroundScheduler()
    sched.add_job(my_job, 'cron', second=par_str, minute='30-59', hour='9')
    sched.add_job(my_job, 'cron', second=par_str, minute='0-29', hour='11')
    sched.add_job(my_job, 'cron', second=par_str, minute='*', hour='10,13,14')
    #sched.add_job(my_job, 'cron', second=par_str, minute='*', hour='*')
    print("开始抓取")
    sched.start()

    # 第五部分，画图



    linelist = []
    animlist = []
    for index in range(len(rectlist)):
        fig = plt.figure(rectlist[index].name)

        line, = plt.plot(range(len(datalist[index])), datalist[index], 'o', markersize=2, label=rectlist[index].name)
        linelist.append(line)

        step_int = float(step)
        plt.xlim(0, 3600*4 / step_int)
        plt.ylim(ylimit_min, ylimit_max)
        plt.legend()
        plt.grid(True)  # 添加网格

        anim = animation.FuncAnimation(fig, animate, fargs=(index, line), interval=step_int * 1000)
        animlist.append(anim)

    plt.show()

    while(1):
        input()

def animate(i, index,line):
    lock.acquire()
    line.set_data(range(len(datalist[index])), datalist[index])
    lock.release()
    return line


#这个类没有实际功能，作为结构体存储数据
class RectItem:
    def __init__(self):
        self.name = ''  # 名字
        self.topleftwidth = 0  # 左上点横坐标
        self.topleftheight = 0  # 左上点纵坐标
        self.bottomrightwidth = 0  # 右下点横坐标
        self.bottomrightheight = 0  # 右下点纵坐标
        self.count = 0

    def __str__(self):
        return self.name + ":左上点 " + str(self.topleftwidth) + " " + str(self.topleftheight) + " 右下点 " + str(self.bottomrightwidth) + " " + str(self.bottomrightheight)

def my_job():
    global tolal_num, right_num
    wb = load_workbook(res_path)
    ws = wb.active
    result_list = [datetime.datetime.now().strftime('%H:%M:%S'), ]
    lock.acquire()
    for index in range(len(rectlist)):
        tolal_num = tolal_num + 1
        im = PIL.ImageGrab.grab((rectlist[index].topleftwidth, rectlist[index].topleftheight, rectlist[index].bottomrightwidth, rectlist[index].bottomrightheight))
        #im = im.resize((im.size[0] * 18, im.size[1] * 18), PIL.Image.ANTIALIAS)
        tmpstr = pytesseract.image_to_string(im, lang="num", config="-psm 8")
        xls_result = ''
        plot_result = 0
        try:
            result = float(tmpstr)
            xls_result = result
            plot_result = result
            pre_date[index] = result
            right_num = right_num + 1
        except Exception as e:
            print(e)
            xls_result = tmpstr
            plot_result = pre_date[index]
        finally:
            result_list.append(xls_result)
            datalist[index].append(plot_result)
    lock.release()
    ws.append(result_list)
    outstr = ''
    for tmp in result_list:
        outstr = outstr + ' ' + str(tmp)
    print('插入行'+outstr)
    print('当前识别率:'+str(right_num/tolal_num))
    wb.save(res_path)

def test_job():
    print("测试任务执行")

if __name__ == '__main__':
    # 下面声明一些全局变量
    # 这个是创建的记录文件的路径
    res_path = 'result/' + datetime.datetime.now().strftime('%Y-%m-%d') + '.xlsx'
    # 这个是操作datalist的线程锁
    lock = threading.Lock()
    # 观测点列表
    rectlist = []
    # 数据二维列表
    datalist = []
    # 前值列表
    pre_date = []
    # 总数据量
    tolal_num = 0
    # 正确的数据量
    right_num = 0
    # 启动主程序
    main()
